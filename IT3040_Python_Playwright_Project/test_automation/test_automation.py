"""
IT3040 Assignment 1 - Option 1
Chat Sinhala Transliteration Accuracy Testing

This script reads test cases from an Excel file, opens the PixelsSuite
Chat Translator page using Playwright, enters each Singlish input, captures
the actual Sinhala output, and writes Actual output + Status back to Excel.

Run example:
python test_automation.py --excel "Assignment 1 - Test cases.xlsx" --url "https://www.pixelssuite.com/chat-translator" --wait-ms 5000 --type-delay-ms 80 --slow-mo-ms 200 --save-every 1 --keep-open
"""

import argparse
import os
import sys
import time
from typing import Optional, Tuple, List

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


# ----------------------------
# Helper functions
# ----------------------------

SINHALA_START = "\u0D80"
SINHALA_END = "\u0DFF"


def has_sinhala(text: str) -> bool:
    """Return True if a string contains Sinhala Unicode characters."""
    return any(SINHALA_START <= ch <= SINHALA_END for ch in text or "")


def clean_text(text: Optional[str]) -> str:
    """Normalize captured web text for comparison/writing."""
    if text is None:
        return ""
    text = str(text).replace("\r\n", "\n").replace("\r", "\n")
    # Remove invisible zero width characters that sometimes appear in copied output.
    text = text.replace("\u200b", "").replace("\ufeff", "")
    return text.strip()


def compare_text(actual: str, expected: str) -> str:
    """
    Assignment status rule:
    Pass = actual output matches expected output
    Fail = actual output does not match expected output
    """
    return "Pass" if clean_text(actual) == clean_text(expected) else "Fail"


def find_header_columns(ws) -> dict:
    """Find required Excel headers and return a header-name to column-number map."""
    header_aliases = {
        "TC ID": ["TC ID", "Test Case ID"],
        "Input length type": ["Input length type"],
        "Input": ["Input"],
        "Expected output": ["Expected output"],
        "Actual output": ["Actual output"],
        "Status": ["Status"],
    }
    required = {name: None for name in header_aliases}
    alias_to_required = {
        alias.strip().lower(): name
        for name, aliases in header_aliases.items()
        for alias in aliases
    }

    for cell in ws[1]:
        value = str(cell.value).strip() if cell.value is not None else ""
        required_name = alias_to_required.get(value.lower())
        if required_name:
            required[required_name] = cell.column

    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(
            "Missing required Excel columns: "
            + ", ".join(missing)
            + "\nRequired columns: TC ID/Test Case ID, Input length type, Input, Expected output, Actual output, Status"
        )

    return required


def load_test_rows(excel_path: str) -> Tuple[object, object, dict, List[int]]:
    """Open workbook and return rows that contain test case data."""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)
    if "Assignment 1 - Test cases" in wb.sheetnames:
        ws = wb["Assignment 1 - Test cases"]
    else:
        ws = wb.active

    cols = find_header_columns(ws)

    test_rows = []
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=cols["TC ID"]).value
        input_text = ws.cell(row=row, column=cols["Input"]).value
        expected = ws.cell(row=row, column=cols["Expected output"]).value
        if tc_id and input_text is not None and expected is not None:
            test_rows.append(row)

    return wb, ws, cols, test_rows


def try_click_text(page, texts: List[str]) -> None:
    """Try to click a visible tab/button by text. Fails silently if not found."""
    for text in texts:
        try:
            locator = page.get_by_text(text, exact=False).first
            if locator.count() > 0 and locator.is_visible(timeout=1000):
                locator.click(timeout=2000)
                page.wait_for_timeout(700)
                return
        except Exception:
            pass


def choose_input_locator(page):
    """
    Select the input area.
    The website UI may change, so this function tries common textbox patterns.
    """
    preferred = [
        "textarea[placeholder*='English']",
        "textarea[placeholder*='Type']",
        "textarea",
    ]
    for selector in preferred:
        try:
            item = page.locator(selector).first
            if item.count() > 0 and item.is_visible(timeout=500) and item.is_enabled(timeout=500):
                return item
        except Exception:
            pass

    candidates = [
        "textarea",
        "input[type='text']",
        "[contenteditable='true']",
        "[role='textbox']",
    ]

    visible = []
    for selector in candidates:
        try:
            locs = page.locator(selector)
            count = locs.count()
            for i in range(count):
                item = locs.nth(i)
                try:
                    if item.is_visible(timeout=500) and item.is_enabled(timeout=500):
                        box = item.bounding_box()
                        area = (box["width"] * box["height"]) if box else 0
                        visible.append((area, selector, i, item))
                except Exception:
                    continue
        except Exception:
            continue

    if not visible:
        raise RuntimeError("Could not find an input textbox/textarea on the page.")

    # Usually the largest visible textbox is the Singlish input box.
    visible.sort(key=lambda x: x[0], reverse=True)
    return visible[0][3]


def click_transliterate_button(page) -> None:
    """Click the page action button that generates the Sinhala output."""
    button_names = ["Transliterate", "Translate", "Convert", "Generate"]
    for name in button_names:
        try:
            button = page.get_by_role("button", name=name, exact=True).first
            if button.count() > 0 and button.is_visible(timeout=700) and button.is_enabled(timeout=700):
                button.click(timeout=3000)
                return
        except Exception:
            pass

    for name in button_names:
        try:
            button = page.get_by_text(name, exact=False).first
            if button.count() > 0 and button.is_visible(timeout=700):
                button.click(timeout=3000)
                return
        except Exception:
            pass


def click_clear_button(page) -> None:
    """Use the page's Clear button so the app resets its own input/output state."""
    try:
        buttons = page.locator("button")
        for i in range(buttons.count()):
            button = buttons.nth(i)
            if button.is_visible(timeout=300) and "Clear" in button.inner_text(timeout=300):
                button.click(timeout=3000)
                page.wait_for_function(
                    """() => Array.from(document.querySelectorAll('textarea'))
                        .every(box => !box.value || box.value.trim().length === 0)""",
                    timeout=5000
                )
                return
    except Exception:
        pass

    try:
        button = page.get_by_text("Clear", exact=False).first
        if button.count() > 0 and button.is_visible(timeout=700):
            button.click(timeout=3000)
            page.wait_for_function(
                """() => Array.from(document.querySelectorAll('textarea'))
                    .every(box => !box.value || box.value.trim().length === 0)""",
                timeout=5000
            )
    except Exception:
        pass


def clear_output_box(page) -> None:
    """Clear any previous generated output before running the next test case."""
    try:
        page.evaluate(
            """() => {
                const boxes = Array.from(document.querySelectorAll('textarea'));
                const output = boxes.find(box => /Sinhala|appear/i.test(box.getAttribute('placeholder') || '')) || boxes[1];
                if (output) {
                    output.value = '';
                    output.dispatchEvent(new Event('input', { bubbles: true }));
                    output.dispatchEvent(new Event('change', { bubbles: true }));
                }
            }"""
        )
    except Exception:
        pass


def get_value_or_text(locator) -> str:
    """Read value from input/textarea or inner text from other elements."""
    try:
        value = locator.input_value(timeout=800)
        if value is not None:
            return clean_text(value)
    except Exception:
        pass
    try:
        return clean_text(locator.inner_text(timeout=800))
    except Exception:
        return ""


def capture_output(page, input_text: str, expected_text: str) -> str:
    """
    Capture the Sinhala output from the page.
    It first checks textareas/inputs, then visible Sinhala text blocks.
    """
    # PixelsSuite currently uses the second textarea as the Sinhala output box.
    output_selectors = [
        "textarea[placeholder*='Sinhala']",
        "textarea[placeholder*='appear']",
    ]
    for selector in output_selectors:
        try:
            item = page.locator(selector).first
            if item.count() > 0 and item.is_visible(timeout=500):
                txt = get_value_or_text(item)
                if txt:
                    return txt
        except Exception:
            pass

    try:
        textareas = page.locator("textarea")
        if textareas.count() >= 2 and textareas.nth(1).is_visible(timeout=500):
            txt = get_value_or_text(textareas.nth(1))
            if txt:
                return txt
    except Exception:
        pass

    # 1) Common case: two textareas/input boxes. First is input, second is output.
    controls = []
    for selector in ["textarea", "input[type='text']", "[contenteditable='true']", "[role='textbox']"]:
        try:
            locs = page.locator(selector)
            for i in range(locs.count()):
                item = locs.nth(i)
                try:
                    if item.is_visible(timeout=500):
                        txt = get_value_or_text(item)
                        if txt and clean_text(txt) != clean_text(input_text):
                            controls.append(txt)
                except Exception:
                    pass
        except Exception:
            pass

    # Prefer candidates that contain Sinhala.
    sinhala_controls = [x for x in controls if has_sinhala(x)]
    if sinhala_controls:
        # Pick the longest Sinhala candidate.
        return max(sinhala_controls, key=len)

    # 2) Check visible page text and find Sinhala-heavy blocks.
    try:
        blocks = page.locator("body *").evaluate_all(
            """els => els
                .filter(e => {
                    const style = window.getComputedStyle(e);
                    const rect = e.getBoundingClientRect();
                    return style && style.visibility !== 'hidden' &&
                           style.display !== 'none' &&
                           rect.width > 20 && rect.height > 10;
                })
                .map(e => (e.innerText || e.textContent || '').trim())
                .filter(t => t.length > 0)
            """
        )
        blocks = [clean_text(b) for b in blocks if clean_text(b)]
        sinhala_blocks = [b for b in blocks if has_sinhala(b) and clean_text(input_text) not in b]
        if sinhala_blocks:
            # Avoid giant body copy; pick a sensible block closest to expected size.
            expected_len = max(len(clean_text(expected_text)), 1)
            sinhala_blocks.sort(key=lambda x: (abs(len(x) - expected_len), -len(x)))
            return sinhala_blocks[0]
    except Exception:
        pass

    return ""


def run_one_test(page, input_text: str, expected_text: str, type_delay_ms: int, wait_ms: int) -> str:
    """Enter one test input and return captured actual output."""
    click_clear_button(page)
    input_box = choose_input_locator(page)

    try:
        input_box.click(timeout=5000)
        input_box.fill("", timeout=3000)
        input_box.type(str(input_text), delay=type_delay_ms)
    except Exception:
        # Fallback keyboard typing
        input_box.click(timeout=5000)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
        page.keyboard.type(str(input_text), delay=type_delay_ms)

    try:
        current_input = clean_text(input_box.input_value(timeout=1000))
        if current_input != clean_text(input_text):
            input_box.click(timeout=5000)
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
            page.keyboard.type(str(input_text), delay=type_delay_ms)
    except Exception:
        pass

    page.wait_for_timeout(1000)
    previous_output = ""
    try:
        previous_output = page.evaluate(
            """() => {
                const boxes = Array.from(document.querySelectorAll('textarea'));
                const output = boxes.find(box => /Sinhala|appear/i.test(box.getAttribute('placeholder') || '')) || boxes[1];
                return output && output.value ? output.value.trim() : '';
            }"""
        )
    except Exception:
        pass

    click_transliterate_button(page)
    try:
        page.wait_for_function(
            """previous => {
                const boxes = Array.from(document.querySelectorAll('textarea'));
                const output = boxes.find(box => /Sinhala|appear/i.test(box.getAttribute('placeholder') || '')) || boxes[1];
                const value = output && output.value ? output.value.trim() : '';
                return value.length > 0 && value !== previous;
            }""",
            arg=previous_output,
            timeout=max(wait_ms, 45000)
        )
    except PlaywrightTimeoutError:
        actual = capture_output(page, input_text, expected_text)
        if actual:
            return actual
        raise RuntimeError("Actual output was not generated within the wait time.")
    return capture_output(page, input_text, expected_text)


# ----------------------------
# Main automation
# ----------------------------

def main():
    parser = argparse.ArgumentParser(description="IT3040 Chat Sinhala transliteration automation")
    parser.add_argument("--excel", required=True, help="Path to Assignment 1 - Test cases.xlsx")
    parser.add_argument("--url", required=True, help="URL of the chat translator")
    parser.add_argument("--wait-ms", type=int, default=5000, help="Wait time after typing input")
    parser.add_argument("--type-delay-ms", type=int, default=80, help="Delay between typed characters")
    parser.add_argument("--slow-mo-ms", type=int, default=200, help="Slow motion delay for Playwright")
    parser.add_argument("--save-every", type=int, default=1, help="Save Excel after every N test cases")
    parser.add_argument("--keep-open", action="store_true", help="Keep browser open at the end")
    parser.add_argument("--headless", action="store_true", help="Run browser headless")
    args = parser.parse_args()

    excel_path = os.path.abspath(args.excel)

    try:
        wb, ws, cols, test_rows = load_test_rows(excel_path)
    except Exception as e:
        print(f"[ERROR] Excel loading failed: {e}")
        sys.exit(1)

    print(f"[INFO] Excel file: {excel_path}")
    print(f"[INFO] Test cases found: {len(test_rows)}")
    print("[INFO] Keep the Excel file closed while this script is running.")

    processed = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=args.headless,
            slow_mo=args.slow_mo_ms
        )
        context = browser.new_context(viewport={"width": 1366, "height": 768})
        page = context.new_page()

        try:
            page.goto(args.url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(3000)
        except PlaywrightTimeoutError:
            print("[WARN] Page load timeout. Continuing with current page state.")

        # Try to select Chat Sinhala mode/function if visible.
        try_click_text(page, ["Chat Sinhala", "Chat", "Sinhala", "Singlish"])

        for row in test_rows:
            tc_id = ws.cell(row=row, column=cols["TC ID"]).value
            input_text = ws.cell(row=row, column=cols["Input"]).value
            expected = ws.cell(row=row, column=cols["Expected output"]).value

            print(f"\n[RUN] {tc_id}")
            print(f"Input: {input_text}")

            try:
                actual = run_one_test(
                    page=page,
                    input_text=str(input_text),
                    expected_text=str(expected),
                    type_delay_ms=args.type_delay_ms,
                    wait_ms=args.wait_ms
                )
                status = compare_text(actual, str(expected))

                ws.cell(row=row, column=cols["Actual output"]).value = actual
                ws.cell(row=row, column=cols["Status"]).value = status

                print(f"Actual: {actual}")
                print(f"Status: {status}")

            except Exception as e:
                error_message = f"Automation Error: {type(e).__name__}: {e}"
                ws.cell(row=row, column=cols["Actual output"]).value = error_message
                ws.cell(row=row, column=cols["Status"]).value = "Error"
                print(f"[ERROR] {tc_id}: {error_message}")

            processed += 1
            if args.save_every > 0 and processed % args.save_every == 0:
                wb.save(excel_path)
                print(f"[SAVE] Excel updated after {processed} test case(s).")

        wb.save(excel_path)
        print("\n[DONE] All test cases processed.")
        print("[DONE] Excel Actual output and Status columns were updated.")

        if args.keep_open:
            print("[INFO] Browser kept open. Press ENTER here to close it.")
            try:
                input()
            except KeyboardInterrupt:
                pass

        context.close()
        browser.close()


if __name__ == "__main__":
    main()
